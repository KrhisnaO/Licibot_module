from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required
from .models import CustomUser, Licitacion, Preguntasbbdd, Respuesta, ErrorHistory
from .forms import LoginForm, CreateUserForm, LicitacionForm, PreguntasForm, SubirArchivoForm, CustomPasswordResetForm, ValidarIDLicitacionForm
import requests
from django.db.models import Q
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile

## MANEJO DE APIS ##
from .utils import obtener_licitaciones, subir_chatpdf, preguntar_chatpdf

## MANEJO DE ERRORES EN EL SISTEMA ##
from .middleware import ErrorLoggingMiddleware
from datetime import datetime
import traceback

## EXCEL ##
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList


## PAGINA HOME ##########################################################
def home(request):
    lici_count = Licitacion.objects.count()
    preg_count = Preguntasbbdd.objects.count()
    nolici_count = Licitacion.objects.filter(Q(archivoLicitacion__isnull=True) | Q(archivoLicitacion="")).count()
    licidoc_count = Licitacion.objects.filter(~Q(archivoLicitacion__isnull=True) & ~Q(archivoLicitacion="")).count()

    # Verifica la pertenencia a grupos
    is_vendedor = request.user.groups.filter(name='VENDEDOR').exists()
    is_gerente = request.user.groups.filter(name='GERENTE').exists()

    context = {
        'lici_count': lici_count,
        'preg_count': preg_count,
        'nolici_count': nolici_count,
        'licidoc_count': licidoc_count,
        'is_vendedor': is_vendedor,
        'is_gerente': is_gerente,
    }

    return render(request, 'core/home.html', context)

##### INGRESO DE SESION ###################################################
def login_view(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user is not None:
                if user.is_active:
                    login(request, user)
                    if user.is_superuser:
                        return redirect('administrador')
                    elif user.groups.filter(name='VENDEDOR').exists():
                        return redirect('administrador')
                    elif user.groups.filter(name='GERENTE').exists():
                        return redirect('administrador')
                    else:
                        return redirect('home')
                else:
                    mesg = "Usuario deshabilitado, en caso de error contactarse con el administrador."
                    return render(request, 'core/ingreso.html', {'form': form, 'mesg': mesg})
            else:
                mesg = "Usuario o contraseña incorrectos."
                return render(request, 'core/ingreso.html', {'form': form, 'mesg': mesg})
    else:
        form = LoginForm()
    return render(request, 'core/ingreso.html', {'form': form})


## CAMBIAR CONTRASEÑA ##################################################
def recuperar_pass(request):
    if request.method == 'POST':
        form = CustomPasswordResetForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            new_password = form.cleaned_data['new_password']

            User = get_user_model()
            user = User.objects.get(email=email)

            user.set_password(new_password)
            user.save()
            return redirect('ingreso') 
    else:
        form = CustomPasswordResetForm()

    return render(request, 'core/recuperar_pass.html', {'form': form})


## CREADOR DE USUARIOS ###############################################################
@login_required
def crear_usuario(request):
    if request.method == 'POST':
        form = CreateUserForm(request.POST)
        if form.is_valid():
            rut = form.cleaned_data['rut']
            if CustomUser.objects.filter(rut=rut).exists():
                descripcion = f"Error al crear usuario: Ya existe un usuario con el RUT {rut}."
                ErrorHistory.objects.create(
                    tipo_vista='crear_usuario',
                    descripcion=descripcion
                )
                messages.error(request, 'Ya existe un usuario con este RUT.')
            else:
                user = form.save()
                messages.success(request, 'El usuario se ha creado correctamente.')
                return redirect('crear_usuario')
        else:
            descripcion = "Error al crear usuario: Datos inválidos en el formulario."
            ErrorHistory.objects.create(
                tipo_vista='crear_usuario',
                descripcion=descripcion
            )
            messages.error(request, 'Hubo un problema al crear el usuario. Contacte al administrador.')
    else:
        form = CreateUserForm()

    return render(request, 'core/crear_usuario.html', {'form': form})

## HISTORIAL DE USUARIO #########################################################
@login_required
def historial_usu(request):
    usuarios = CustomUser.objects.all()

    # Limpiar los mensajes después de recuperarlos
    storage = messages.get_messages(request)
    storage.used = False

    return render(request, 'core/historial_usu.html', {'usuarios': usuarios})

#### EDITAR USUARIO ################################################################@login_required
def editar_usuario(request, user_id):
    usuario = get_object_or_404(CustomUser, id=user_id)
    
    if request.method == 'POST':
        form = CreateUserForm(request.POST, instance=usuario)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Se ha actualizado la información exitosamente.')
                return redirect('historial_usu')
            except Exception as e:
                tipo_vista = 'editar_usuario'
                descripcion = f"Error al actualizar usuario (ID: {user_id}): {str(e)}"
                registrar_error(request, tipo_vista, descripcion)
                messages.error(request, 'Hubo un problema al actualizar la información del usuario. Contacte al administrador.')
    else:
        form = CreateUserForm(instance=usuario)

    # Limpiar los mensajes después de recuperarlos
    storage = messages.get_messages(request)
    storage.used = False
    
    return render(request, 'core/editar_usuario.html', {'form': form})

# CERRAR SESION ######################################################################
@login_required
def cerrar_sesion(request):
    logout(request)
    return redirect('home')

# LOGIN ADMINISTRADR ################################################################
@login_required
def administrador(request):
    return render(request, 'core/administrador.html')

# VISTA DE GERENTE HISTORIAL CON USUARIOS ############################################
@login_required
def historial_lici_con_usuario(request):
    licitaciones_guardadas = Licitacion.objects.select_related('subido_por').all()

    return render(request, 'core/historial_lici_con_usuario.html', {'licitaciones_guardadas': licitaciones_guardadas})


#### MANTENEDOR DE PREGUNTAS #############################################
@login_required
def mantenedor_preguntas(request, action, id):
    data = {"mesg": "", "action": action, "id": id}

    try:
        if action == 'ins':
            if request.method == "POST":
                form = PreguntasForm(request.POST, request.FILES)
                if form.is_valid():
                    try:
                        form.save()
                        data["mesg"] = "¡La Pregunta fue creada correctamente!"
                    except Exception as e:
                        registrar_error(request, 'mantenedor_preguntas', f"Error al crear pregunta: {str(e)}")
                        data["mesg"] = "¡No se pudo crear la pregunta!"
            else:
                form = PreguntasForm()

        elif action == 'upd':
            objeto = Preguntasbbdd.objects.get(idPreguntas=id)
            if request.method == "POST":
                form = PreguntasForm(request.POST, request.FILES, instance=objeto)
                if form.is_valid():
                    try:
                        form.save()
                        data["mesg"] = "¡La Pregunta fue actualizada correctamente!"
                    except Exception as e:
                        registrar_error(request, 'mantenedor_preguntas', f"Error al actualizar pregunta: {str(e)}")
                        data["mesg"] = "¡No se pudo actualizar la pregunta!"
            else:
                form = PreguntasForm(instance=objeto)

        elif action == 'del':
            try:
                Preguntasbbdd.objects.get(idPreguntas=id).delete()
                data["mesg"] = "¡La Pregunta fue eliminada correctamente!"
                return redirect('mantenedor_preguntas', action='ins', id='-1')
            except Exception as e:
                registrar_error(request, 'mantenedor_preguntas', f"Error al eliminar pregunta: {str(e)}")
                data["mesg"] = "¡No se pudo eliminar la pregunta!"

        data["list"] = Preguntasbbdd.objects.all().order_by('idPreguntas')
        data["form"] = form

    except Exception as e:
        registrar_error(request, 'mantenedor_preguntas', f"Error en mantenedor de preguntas: {str(e)}")
        messages.error(request, 'Ha ocurrido un error al procesar la solicitud.')

    # Limpiar los mensajes después de recuperarlos
    storage = messages.get_messages(request)
    storage.used = False
    return render(request, "core/mantenedor_preguntas.html", data)

##############################################################################################
##############################################################################################
## COMIENZO LICITACIONES ####

## VALIDAR ID LICITACION ###
@login_required
def validar_licitacion(request):
    if request.method == 'POST':
        form = ValidarIDLicitacionForm(request.POST)
        if form.is_valid():
            id_licitacion = form.cleaned_data['idLicitacion']
            licitaciones = obtener_licitaciones(filtro_id=id_licitacion)
            
            if licitaciones:
                licitacion_data = licitaciones[0]
                licitacion, created = Licitacion.objects.update_or_create(
                    idLicitacion=licitacion_data['CodigoExterno'],
                    defaults={
                        'nombreLicitacion': licitacion_data.get('Nombre', ''),
                        'descripcionLicitacion': licitacion_data.get('Descripcion', ''),
                        'fechaCierre': licitacion_data.get('FechaCierre', None),
                        'estado': licitacion_data.get('Estado', ''),
                        'nombreOrganismo': licitacion_data.get('NombreOrganismo', ''),
                        'diasCierreLicitacion': licitacion_data.get('DiasCierreLicitacion', None)  
                    }
                )
                
                if created:
                    messages.success(request, f'ID de Licitación {id_licitacion} válido. Ahora puede subir el archivo PDF.')
                else:
                    messages.info(request, f'ID de Licitación {id_licitacion} ya existe. Puede actualizar el archivo PDF.')

                request.session['idLicitacion'] = id_licitacion
                request.session['licitacion_validada'] = True
                return redirect('subir_archivo_lici', id=id_licitacion)
            else:
                messages.error(request, f'ID de Licitación {id_licitacion} no encontrado en Mercado Público. Por favor, verifique el ID.')
                request.session['licitacion_validada'] = False
        else:
            messages.error(request, 'Error en el formulario. Por favor, revise los datos ingresados.')
            request.session['licitacion_validada'] = False
    else:
        form = ValidarIDLicitacionForm()
        request.session['licitacion_validada'] = False

    return render(request, 'core/validar_licitacion.html', {'form': form})


## SUBIR ARCHIVO LICITACION ###
@login_required
def subir_archivo_lici(request, id):
    licitacion = get_object_or_404(Licitacion, idLicitacion=id)

    if request.method == 'POST':
        form = SubirArchivoForm(request.POST, request.FILES, instance=licitacion)

        if form.is_valid():
            archivo = form.cleaned_data.get('archivoLicitacion')
            if archivo:
                if not archivo.name.endswith('.pdf'):
                    messages.error(request, 'Solo se permiten archivos PDF.')
                else:
                    file_name = default_storage.save(archivo.name, ContentFile(archivo.read()))
                    file_path = default_storage.path(file_name)

                    source_id = subir_chatpdf(file_path)
                    if source_id:
                        try:
                            id_pregunta = Preguntasbbdd.objects.get(nombrePregunta="ID de la licitación").nombrePregunta
                            respuesta_id = preguntar_chatpdf(source_id, id_pregunta)

                            if respuesta_id and str(licitacion.idLicitacion) in respuesta_id:
                                licitacion.archivoLicitacion = file_name
                                licitacion.save()
                                messages.success(request, 'Archivo validado correctamente por ID de licitación.')
                                request.session['file_name'] = file_name
                                return redirect('seleccionar_preguntas', licitacion_id=licitacion.idLicitacion)
                            else:
                                messages.error(request, 'El archivo no pertenece a la licitación. Ningún parámetro coincide.')
                        except Preguntasbbdd.DoesNotExist as e:
                            messages.error(request, f'No se encontró la pregunta necesaria en la base de datos: {str(e)}')
                    else:
                        messages.error(request, 'Error al subir el archivo PDF o no se validó correctamente.')

            else:
                messages.error(request, 'No se ha seleccionado ningún archivo para subir.')
        else:
            messages.error(request, 'Error en el formulario.')
    else:
        form = SubirArchivoForm(instance=licitacion)

    return render(request, 'core/subir_archivo_lici.html', {'form': form, 'id_licitacion': id})


@login_required
def seleccionar_preguntas(request, licitacion_id):
    licitacion = get_object_or_404(Licitacion, idLicitacion=licitacion_id)
    preguntas = Preguntasbbdd.objects.all()

    if request.method == 'POST':
        preguntas_seleccionadas = request.POST.getlist('preguntas')
        if preguntas_seleccionadas:
            for pregunta_id in preguntas_seleccionadas:
                pregunta = get_object_or_404(Preguntasbbdd, idPreguntas=pregunta_id)
                
                if licitacion.archivoLicitacion and licitacion.archivoLicitacion.path:
                    try:
                        respuesta_texto = preguntar_chatpdf(licitacion.archivoLicitacion.path, pregunta.nombrePregunta)
                        if respuesta_texto is not None:
                            respuesta = Respuesta.objects.create(
                                licitacion=licitacion,
                                pregunta=pregunta,
                                textoRespuesta=respuesta_texto
                            )
                        else:
                            messages.error(request, f'No se pudo obtener respuesta para la pregunta "{pregunta.nombrePregunta}"')
                    except Exception as e:
                        messages.error(request, f'Error al hacer la pregunta: {str(e)}')
                else:
                    messages.error(request, 'La licitación no tiene un archivo asociado.')
            
            return redirect('leer_pdf', id=licitacion_id)
        else:
            messages.error(request, 'Debe seleccionar al menos una pregunta.')

    return render(request, 'core/filtro_pregun.html', {'preguntas': preguntas, 'licitacion': licitacion})

@login_required
def leer_pdf(request, id):
    licitacion = get_object_or_404(Licitacion, idLicitacion=id)
    preguntas_seleccionadas = request.session.get('preguntas_seleccionadas', [])

    preguntas_respuestas = []
    for pregunta_id in preguntas_seleccionadas:
        pregunta = get_object_or_404(Preguntasbbdd, idPreguntas=pregunta_id)
        respuesta = Respuesta.objects.filter(licitacion=licitacion, pregunta=pregunta).first()
        respuesta_texto = respuesta.textoRespuesta if respuesta else 'No hay respuesta'

        preguntas_respuestas.append({
            'pregunta': pregunta,
            'respuesta': respuesta_texto
        })

    return render(request, 'core/leer_pdf.html', {
        'licitacion': licitacion,
        'preguntas_respuestas': preguntas_respuestas,
    })


@login_required
def guardar_respuestas(request, id):
    if request.method == 'POST':
        licitacion = get_object_or_404(Licitacion, idLicitacion=id)
        preguntas_seleccionadas = request.session.get('preguntas_seleccionadas', [])

        for pregunta_id in preguntas_seleccionadas:
            pregunta = get_object_or_404(Preguntasbbdd, idPreguntas=pregunta_id)
            
            if licitacion.archivoLicitacion and licitacion.archivoLicitacion.path:
                respuesta_texto = preguntar_chatpdf(licitacion.archivoLicitacion.path, pregunta.nombrePregunta)
                if respuesta_texto is not None:
                    respuesta, created = Respuesta.objects.get_or_create(
                        licitacion=licitacion,
                        pregunta=pregunta,
                        defaults={'textoRespuesta': respuesta_texto}
                    )
                    if created:
                        messages.success(request, f'Respuesta para la pregunta "{pregunta.nombrePregunta}" guardada correctamente.')
                else:
                    messages.error(request, f'No se pudo obtener respuesta para la pregunta "{pregunta.nombrePregunta}".')
            else:
                messages.error(request, 'La licitación no tiene un archivo asociado.')
        del request.session['preguntas_seleccionadas']

    return redirect('leer_pdf', id=id)



# HISTORIAL LICITACIONES ##################################
@login_required
def historial_lici(request):
    licitaciones = Licitacion.objects.all()
    return render(request, 'core/historial_lici.html', {'licitaciones': licitaciones})


### VER TODA LA INFORMACION DE LA LICITACION + RESPUESTAS CHATPDF
@login_required
def ver_info_licitacion(request, id):
    licitacion = get_object_or_404(Licitacion, idLicitacion=id)
    respuestas = Respuesta.objects.filter(licitacion=licitacion).select_related('pregunta')

    preguntas_respuestas = []
    for respuesta in respuestas:
        preguntas_respuestas.append({
            'pregunta': respuesta.pregunta.nombrePregunta,
            'respuesta': respuesta.textoRespuesta,
        })

    return render(request, 'core/ver_info_licitacion.html', {
        'licitacion': licitacion,
        'preguntas_respuestas': preguntas_respuestas,
    })


#############################################################################
#############################################################################
### HISTORIAL DE ERRORES ##
@login_required
def registrar_error(request, tipo_vista, descripcion):
    error = ErrorHistory(
        tipo_vista=tipo_vista,
        descripcion=descripcion,
    )
    error.save()
    return redirect('home')

@login_required
def historial_errores(request):
    errores = ErrorHistory.objects.all().order_by('-fecha')
    return render(request, 'core/historial_errores.html', {'errores': errores})

############################################################################################
############################################################################################
## DESCARGA DE ARCHIVOS EXCEL ##

## ARCHIVO EXCEL HISTORIAL DE USUARIO ##
def desc_user_excel(request):
    usuarios = CustomUser.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Usuarios' 

    headers = ['RUT', 'Nombre', 'Apellido', 'Correo electrónico', 'Tipo de Usuario', 'Estado']

    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_num, value=header)

    for row_num, usuario in enumerate(usuarios, start=2):
        ws.cell(row=row_num, column=1, value=usuario.rut)
        ws.cell(row=row_num, column=2, value=usuario.first_name)
        ws.cell(row=row_num, column=3, value=usuario.last_name)
        ws.cell(row=row_num, column=4, value=usuario.email)
        tipo_usuario = next((group.name for group in usuario.groups.all()), 'Otro')
        ws.cell(row=row_num, column=5, value=tipo_usuario)
        estado_usuario = 'Activo' if usuario.is_active else 'Inactivo'
        ws.cell(row=row_num, column=6, value=estado_usuario)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=usuarios.xlsx'

    wb.save(response)

    return response

## ARCHIVO EXCEL HISTORIAL DE LICITACIONES ##
def desc_lici_excel(request):
    licitaciones = Licitacion.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Licitaciones' 

    headers = ['ID de Licitación', 'Nombre de Licitación', 'Archivo de Licitación']

    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_num, value=header)

    total_licitaciones = 0
    licitaciones_con_archivo = 0
    licitaciones_sin_archivo = 0

    for row_num, licitacion in enumerate(licitaciones, start=2):
        total_licitaciones += 1
        ws.cell(row=row_num, column=1, value=licitacion.idLicitacion)
        ws.cell(row=row_num, column=2, value=licitacion.nombreLicitacion)
        if licitacion.archivoLicitacion:
            licitaciones_con_archivo += 1
            ws.cell(row=row_num, column=4, value="Sí")
        else:
            licitaciones_sin_archivo += 1
            ws.cell(row=row_num, column=4, value="No")

    porcentaje_con_archivo = (licitaciones_con_archivo / total_licitaciones) * 100 if total_licitaciones > 0 else 0
    porcentaje_sin_archivo = (licitaciones_sin_archivo / total_licitaciones) * 100 if total_licitaciones > 0 else 0

    ws.cell(row=row_num + 2, column=1, value="Resumen de Licitaciones")
    ws.cell(row=row_num + 3, column=1, value="Total de Licitaciones")
    ws.cell(row=row_num + 3, column=2, value=total_licitaciones)
    ws.cell(row=row_num + 4, column=1, value="Licitaciones con Archivo")
    ws.cell(row=row_num + 4, column=2, value=licitaciones_con_archivo)
    ws.cell(row=row_num + 4, column=3, value=f"{porcentaje_con_archivo:.2f}%")
    ws.cell(row=row_num + 5, column=1, value="Licitaciones sin Archivo")
    ws.cell(row=row_num + 5, column=2, value=licitaciones_sin_archivo)
    ws.cell(row=row_num + 5, column=3, value=f"{porcentaje_sin_archivo:.2f}%")

    chart = PieChart()
    labels = Reference(ws, min_col=1, min_row=row_num + 3, max_row=row_num + 5)
    data = Reference(ws, min_col=2, min_row=row_num + 3, max_row=row_num + 5)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.title = "Estado de Licitaciones"
    
    data_labels = DataLabelList()
    data_labels.showPercent = True 
    data_labels.showVal = False  
    chart.dataLabels = data_labels

    ws.add_chart(chart, "E1")  

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=licitaciones.xlsx'

    wb.save(response)

    return response

